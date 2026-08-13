"""Pruebas de validación para los fixtures sintéticos y la configuración del proyecto.

Verifica que:
- Los archivos fixture existen y son legibles.
- Los archivos Excel/CSV contienen los encabezados y datos esperados.
- El archivo vacío no contiene filas de datos.
- El archivo con encoding ISO-8859-1 se lee correctamente con charset-normalizer.
- Los módulos placeholder del paquete ``src`` son importables.
"""

from __future__ import annotations

from typing import TYPE_CHECKING

import openpyxl
import pytest

if TYPE_CHECKING:
    from pathlib import Path


class TestFixturesExist:
    """Verifica la existencia y accesibilidad de todos los fixtures."""

    def test_fixtures_dir_exists(self, fixtures_dir: Path) -> None:
        assert fixtures_dir.exists()
        assert fixtures_dir.is_dir()

    def test_sample_exportaciones_exists(self, sample_exportaciones_path: Path) -> None:
        assert sample_exportaciones_path.exists()
        assert sample_exportaciones_path.suffix == ".xlsx"
        assert sample_exportaciones_path.stat().st_size > 0

    def test_sample_importaciones_exists(self, sample_importaciones_path: Path) -> None:
        assert sample_importaciones_path.exists()
        assert sample_importaciones_path.suffix == ".csv"
        assert sample_importaciones_path.stat().st_size > 0

    def test_sample_empty_exists(self, sample_empty_path: Path) -> None:
        assert sample_empty_path.exists()
        assert sample_empty_path.suffix == ".xlsx"
        assert sample_empty_path.stat().st_size > 0

    def test_sample_bad_encoding_exists(self, sample_bad_encoding_path: Path) -> None:
        assert sample_bad_encoding_path.exists()
        assert sample_bad_encoding_path.suffix == ".csv"
        assert sample_bad_encoding_path.stat().st_size > 0


class TestExportacionesFixture:
    """Verifica el contenido del fixture de exportaciones (Excel)."""

    EXPECTED_COLUMNS = [
        "fecha",
        "codigo_nandina",
        "descripcion_nandina",
        "pais_iso",
        "pais_nombre",
        "tipo_operacion",
        "valor_fob_usd",
        "peso_bruto_kg",
        "peso_neto_kg",
        "id_departamento",
        "id_via_transporte",
        "id_aduana",
    ]

    def test_has_correct_headers(self, sample_exportaciones_path: Path) -> None:
        wb = openpyxl.load_workbook(sample_exportaciones_path, read_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        wb.close()
        assert headers == self.EXPECTED_COLUMNS

    def test_has_data_rows(self, sample_exportaciones_path: Path) -> None:
        wb = openpyxl.load_workbook(sample_exportaciones_path, read_only=True)
        ws = wb.active
        row_count = sum(1 for _ in ws.iter_rows(min_row=2)) - 0
        wb.close()
        assert row_count == 5

    def test_all_rows_are_exportacion(self, sample_exportaciones_path: Path) -> None:
        wb = openpyxl.load_workbook(sample_exportaciones_path, read_only=True)
        ws = wb.active
        tipo_col_idx = self.EXPECTED_COLUMNS.index("tipo_operacion")
        for row in ws.iter_rows(min_row=2, values_only=True):
            assert row[tipo_col_idx] == "EXPORTACION"
        wb.close()

    def test_peso_bruto_gte_peso_neto(self, sample_exportaciones_path: Path) -> None:
        """Coherencia física: peso_bruto_kg >= peso_neto_kg."""
        wb = openpyxl.load_workbook(sample_exportaciones_path, read_only=True)
        ws = wb.active
        bruto_idx = self.EXPECTED_COLUMNS.index("peso_bruto_kg")
        neto_idx = self.EXPECTED_COLUMNS.index("peso_neto_kg")
        for row in ws.iter_rows(min_row=2, values_only=True):
            assert row[bruto_idx] >= row[neto_idx], (
                f"peso_bruto_kg ({row[bruto_idx]}) < peso_neto_kg ({row[neto_idx]})"
            )
        wb.close()

    def test_valor_fob_non_negative(self, sample_exportaciones_path: Path) -> None:
        """Integridad de rangos: valor_fob_usd >= 0."""
        wb = openpyxl.load_workbook(sample_exportaciones_path, read_only=True)
        ws = wb.active
        fob_idx = self.EXPECTED_COLUMNS.index("valor_fob_usd")
        for row in ws.iter_rows(min_row=2, values_only=True):
            assert row[fob_idx] >= 0, f"valor_fob_usd negativo: {row[fob_idx]}"
        wb.close()


class TestImportacionesFixture:
    """Verifica el contenido del fixture de importaciones (CSV UTF-8)."""

    def test_is_utf8(self, sample_importaciones_path: Path) -> None:
        content = sample_importaciones_path.read_text(encoding="utf-8")
        assert len(content) > 0

    def test_has_correct_row_count(self, sample_importaciones_path: Path) -> None:
        lines = sample_importaciones_path.read_text(encoding="utf-8").strip().split("\n")
        # First line is header, rest are data
        data_lines = lines[1:]
        assert len(data_lines) == 5

    def test_all_rows_are_importacion(self, sample_importaciones_path: Path) -> None:
        lines = sample_importaciones_path.read_text(encoding="utf-8").strip().split("\n")
        header = lines[0].split(",")
        tipo_idx = header.index("tipo_operacion")
        for line in lines[1:]:
            fields = line.split(",")
            assert fields[tipo_idx] == "IMPORTACION"


class TestEmptyFixture:
    """Verifica que el fixture vacío tiene encabezados pero ninguna fila de datos."""

    def test_has_headers_only(self, sample_empty_path: Path) -> None:
        wb = openpyxl.load_workbook(sample_empty_path, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        # Debe tener exactamente 1 fila (encabezados)
        assert len(rows) == 1

    def test_headers_match_schema(self, sample_empty_path: Path) -> None:
        wb = openpyxl.load_workbook(sample_empty_path, read_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        wb.close()
        assert "fecha" in headers
        assert "codigo_nandina" in headers
        assert "valor_fob_usd" in headers


class TestBadEncodingFixture:
    """Verifica que el fixture con encoding ISO-8859-1 se detecta y lee correctamente."""

    def test_is_not_utf8(self, sample_bad_encoding_path: Path) -> None:
        """El archivo NO debe ser UTF-8 válido si se leen caracteres con acento."""
        raw_bytes = sample_bad_encoding_path.read_bytes()
        # La secuencia de bytes para 'é' en ISO-8859-1 es 0xe9 (un solo byte),
        # mientras que en UTF-8 sería 0xc3 0xa9 (dos bytes).
        # Verificamos que existe un byte > 127 que NO es parte de una secuencia UTF-8 válida.
        with pytest.raises(UnicodeDecodeError):
            raw_bytes.decode("utf-8")

    def test_reads_correctly_as_latin1(self, sample_bad_encoding_path: Path) -> None:
        content = sample_bad_encoding_path.read_text(encoding="iso-8859-1")
        assert "fecha" in content
        assert len(content.strip().split("\n")) == 4  # 1 header + 3 data rows

    def test_charset_normalizer_detects_encoding(self, sample_bad_encoding_path: Path) -> None:
        """charset-normalizer debe detectar el encoding correctamente."""
        from charset_normalizer import from_bytes

        raw_bytes = sample_bad_encoding_path.read_bytes()
        result = from_bytes(raw_bytes).best()
        assert result is not None
        detected_encoding = result.encoding.lower()
        # charset-normalizer puede reportar variantes como 'iso8859-1', 'latin-1',
        # 'cp1252', 'cp1250'. Todos son compatibles para los caracteres usados.
        compatible_encodings = {
            "iso-8859-1", "latin-1", "iso8859-1",
            "cp1252", "windows-1252", "cp1250",
        }
        assert detected_encoding in compatible_encodings, (
            f"Encoding detectado '{detected_encoding}' no es compatible con ISO-8859-1"
        )


class TestSourceModulesImportable:
    """Verifica que los módulos placeholder en src/ son importables."""

    def test_import_src_package(self) -> None:
        import src

        assert hasattr(src, "__version__")

    def test_import_extract(self) -> None:
        import src.extract  # noqa: F401

    def test_import_transform(self) -> None:
        import src.transform  # noqa: F401

    def test_import_load(self) -> None:
        import src.load  # noqa: F401

    def test_import_validate(self) -> None:
        import src.validate  # noqa: F401

    def test_import_config(self) -> None:
        import src.config  # noqa: F401

    def test_import_firestore_models(self) -> None:
        import src.firestore_models  # noqa: F401

    def test_import_firestore_schemas(self) -> None:
        import src.firestore_schemas  # noqa: F401
