"""Módulo base de extracción y lectura de datos del INE Bolivia.

Proporciona funciones para leer archivos Excel (.xlsx) y CSV de comercio exterior
descargados del portal del INE, con soporte para lectura en modo read-only para
archivos grandes, detección de encoding y extracción de metadatos.

Para las operaciones de web scraping y descarga automatizada, véase
el módulo especializado ``src.extract_comercio_exterior``.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd

from src.extract_comercio_exterior import (
    INE_EXPORTACIONES_URL,
    INE_IMPORTACIONES_URL,
    ExtractionMetadata,
    ExtractionSummary,
    ScrapedResource,
    compute_sha256,
    create_resilient_session,
    download_resource,
    extract_comercio_exterior,
    scrape_ine_resources,
)

# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------
_SUPPORTED_EXTENSIONS = {".xlsx", ".xls", ".csv"}


# ---------------------------------------------------------------------------
# Funciones de Lectura y Metadatos de Archivos Locales
# ---------------------------------------------------------------------------
def read_ine_excel(
    filepath: str | Path,
    *,
    max_rows: int | None = None,
    engine: str = "openpyxl",
) -> pd.DataFrame:
    """Lee un archivo Excel del INE Bolivia y retorna un DataFrame.

    Parameters
    ----------
    filepath:
        Ruta al archivo `.xlsx` del INE.
    max_rows:
        Número máximo de filas a leer (``None`` para leer todas).
        Útil para archivos grandes de importaciones (~60 MB, ~400k filas).
    engine:
        Motor de lectura de Excel. Por defecto ``openpyxl``.
        Para archivos muy grandes, considerar ``calamine``.

    Returns
    -------
    pd.DataFrame
        DataFrame con los datos crudos del archivo.

    Raises
    ------
    FileNotFoundError
        Si el archivo no existe.
    ValueError
        Si la extensión del archivo no es soportada.
    """
    path = Path(filepath)
    if not path.exists():
        msg = f"Archivo no encontrado: {path}"
        raise FileNotFoundError(msg)

    if path.suffix.lower() not in _SUPPORTED_EXTENSIONS:
        msg = f"Extensión no soportada: {path.suffix!r}. Soportadas: {_SUPPORTED_EXTENSIONS}"
        raise ValueError(msg)

    if path.suffix.lower() == ".csv":
        return _read_csv_with_encoding(path, max_rows=max_rows)

    return pd.read_excel(
        path,
        engine=engine,
        nrows=max_rows,
        dtype=str,  # Leer todo como string para preservar ceros a la izquierda
    )


def _read_csv_with_encoding(
    filepath: Path,
    *,
    max_rows: int | None = None,
) -> pd.DataFrame:
    """Lee un archivo CSV detectando encoding automáticamente.

    Intenta UTF-8 primero; si falla, usa ``charset_normalizer`` para
    detectar el encoding correcto.
    """
    try:
        return pd.read_csv(filepath, encoding="utf-8", nrows=max_rows, dtype=str)
    except UnicodeDecodeError:
        pass

    from charset_normalizer import from_bytes

    raw = filepath.read_bytes()
    result = from_bytes(raw).best()
    if result is None:
        msg = f"No se pudo detectar el encoding de: {filepath}"
        raise ValueError(msg)

    detected = result.encoding
    return pd.read_csv(filepath, encoding=detected, nrows=max_rows, dtype=str)


def get_excel_metadata(filepath: str | Path) -> dict[str, Any]:
    """Extrae metadatos de un archivo Excel del INE sin cargar todos los datos.

    Parameters
    ----------
    filepath:
        Ruta al archivo `.xlsx`.

    Returns
    -------
    dict
        Diccionario con ``sheet_names``, ``headers``, ``n_columns``,
        ``filename`` y ``file_size_mb``.
    """
    import openpyxl

    path = Path(filepath)
    if not path.exists():
        msg = f"Archivo no encontrado: {path}"
        raise FileNotFoundError(msg)

    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        msg = f"El libro Excel '{path}' no tiene una hoja activa válida."
        raise ValueError(msg)

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    sheet_names = wb.sheetnames
    wb.close()

    return {
        "filename": path.name,
        "file_size_mb": round(path.stat().st_size / (1024 * 1024), 2),
        "sheet_names": sheet_names,
        "active_sheet": ws.title,
        "headers": headers,
        "n_columns": len(headers),
    }


def list_raw_files(
    directory: str | Path,
    *,
    extension: str = ".xlsx",
    exclude_pattern: str = "DICCIONARIO",
) -> list[Path]:
    """Lista archivos de datos en un directorio, excluyendo diccionarios.

    Parameters
    ----------
    directory:
        Ruta al directorio de archivos raw.
    extension:
        Extensión a buscar (por defecto ``.xlsx``).
    exclude_pattern:
        Patrón en el nombre de archivo para excluir (ej: ``DICCIONARIO``).

    Returns
    -------
    list[Path]
        Lista de rutas a archivos de datos ordenados por nombre.
    """
    path = Path(directory)
    if not path.is_dir():
        msg = f"Directorio no encontrado: {path}"
        raise FileNotFoundError(msg)

    return sorted(
        f
        for f in path.glob(f"*{extension}")
        if exclude_pattern.upper() not in f.name.upper()
    )


__all__ = [
    "INE_EXPORTACIONES_URL",
    "INE_IMPORTACIONES_URL",
    "ExtractionMetadata",
    "ExtractionSummary",
    "ScrapedResource",
    "_read_csv_with_encoding",
    "compute_sha256",
    "create_resilient_session",
    "download_resource",
    "extract_comercio_exterior",
    "get_excel_metadata",
    "list_raw_files",
    "read_ine_excel",
    "scrape_ine_resources",
]
